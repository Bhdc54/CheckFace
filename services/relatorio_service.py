import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from repositories.aula_repository import AulaRepository
from repositories.presenca_repository import PresencaRepository


class RelatorioService:

    def __init__(self):
        self.aula_repo = AulaRepository()
        self.presenca_repo = PresencaRepository()

    def gerar_excel(self, aula_id: int) -> str:
        aula = self.aula_repo.buscar_por_id(aula_id)
        presencas = self.presenca_repo.listar_detalhado_por_aula(aula_id)
        todos = self.aula_repo.buscar_alunos_da_turma(aula_id)

        presentes_matriculas = {p[0] for p in presencas}
        presenca_dict = {p[0]: p for p in presencas}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Presenças"
        # Título
        ws.merge_cells("A1:E1")
        ws["A1"] = f"Relatório de Presença — {aula[3]}"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A2"] = f"Professor: {aula[4]}"
        ws["B2"] = f"Início: {aula[2].strftime('%d/%m/%Y %H:%M')}"
        ws["D2"] = f"Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        ws.append([])

        # Cabeçalhos
        headers = ["Matrícula", "Nome", "Status", "Hora de entrada", "Confiança (%)"]
        ws.append(headers)
        for col in range(1, 6):
            cell = ws.cell(row=4, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A73E8")
            cell.alignment = Alignment(horizontal="center")

        # Dados dos alunos
        for i, aluno in enumerate(todos, start=5):
            matricula, nome = aluno
            if matricula in presentes_matriculas:
                p = presenca_dict[matricula]
                status = "Presente"
                hora = str(p[2])
                confianca = f"{round(p[3] * 100)}%" if p[3] else "-"
                cor_status = "1E8449"
            else:
                status = "Ausente"
                hora = "-"
                confianca = "-"
                cor_status = "C0392B"

            ws.cell(i, 1, matricula)
            ws.cell(i, 2, nome)
            cell_status = ws.cell(i, 3, status)
            cell_status.font = Font(bold=True, color=cor_status)
            ws.cell(i, 4, hora)
            ws.cell(i, 5, confianca)

        # Resumo
        linha_resumo = len(todos) + 6
        ws.cell(linha_resumo, 1, "Total presentes:").font = Font(bold=True)
        ws.cell(linha_resumo, 2, f"{len(presencas)} de {len(todos)}")
        ws.cell(linha_resumo + 1, 1, "Percentual:").font = Font(bold=True)
        pct = round((len(presencas) / len(todos)) * 100) if todos else 0
        ws.cell(linha_resumo + 1, 2, f"{pct}%")

        # Largura das colunas
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 14

        pasta = os.path.join(os.path.dirname(__file__), '..', 'relatorios')
        os.makedirs(pasta, exist_ok=True)
        caminho = os.path.abspath(os.path.join(pasta, f"relatorio_aula_{aula_id}.xlsx"))
        wb.save(caminho)
        return caminho